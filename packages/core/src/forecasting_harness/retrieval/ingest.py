from __future__ import annotations

import csv
from email import message_from_binary_file, policy
import json
import plistlib
import re
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

from forecasting_harness.retrieval.registry import CorpusRegistry, parse_published_at


@dataclass(frozen=True)
class IngestedChunk:
    chunk_id: str
    location: str
    content: str


@dataclass(frozen=True)
class IngestedDocument:
    source_id: str
    title: str
    source_type: str
    path: str
    published_at: str | None
    tags: dict[str, str]
    chunks: list[IngestedChunk]


def _sanitize_source_id(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return normalized or "source"


def detect_source_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".md", ".txt"}:
        return "markdown"
    if suffix == ".csv":
        return "csv"
    if suffix == ".json":
        return "json"
    if suffix == ".xlsx":
        return "spreadsheet"
    if suffix in {".html", ".htm"}:
        return "html"
    if suffix in {".mhtml", ".mht", ".eml", ".webarchive"}:
        return "web-archive"
    if suffix == ".pdf":
        return "pdf"
    raise ValueError(f"unsupported source type: {suffix or path.name}")


def _normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _infer_published_at(value: str | None) -> str | None:
    if value is None:
        return None
    candidate = value.strip()
    if not candidate:
        return None
    if len(candidate) >= 10 and re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", candidate):
        candidate = candidate[:10]
    try:
        return parse_published_at(candidate)
    except ValueError:
        return None


def _parse_spreadsheet_chunks(path: Path) -> list[IngestedChunk]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    chunks: list[IngestedChunk] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                cells: list[str] = []
                min_col: int | None = None
                max_col: int | None = None
                row_number: int | None = None
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    text = _normalize_whitespace(str(value))
                    if not text:
                        continue
                    row_number = cell.row if row_number is None else row_number
                    min_col = cell.column if min_col is None else min(min_col, cell.column)
                    max_col = cell.column if max_col is None else max(max_col, cell.column)
                    cells.append(f"{cell.coordinate}={text}")
                if not cells or row_number is None or min_col is None or max_col is None:
                    continue
                start = f"{get_column_letter(min_col)}{row_number}"
                end = f"{get_column_letter(max_col)}{row_number}"
                location = f"sheet:{sheet.title}!{start}:{end}"
                chunks.append(
                    IngestedChunk(
                        chunk_id=str(len(chunks) + 1),
                        location=location,
                        content=" | ".join(cells),
                    )
                )
    finally:
        workbook.close()
    if not chunks:
        raise ValueError(f"xlsx has no extractable text: {path}")
    return chunks


_PUBLISHED_AT_META_KEYS = {
    "article:published_time",
    "date",
    "datepublished",
    "dc.date",
    "dc.date.issued",
    "pubdate",
    "published_time",
}
_TITLE_META_KEYS = {"og:title", "title", "twitter:title"}
_HTML_CHARSET_RE = re.compile(br"charset\s*=\s*['\"]?\s*([A-Za-z0-9._-]+)", re.IGNORECASE)


class _SavedPageParser(HTMLParser):
    _BLOCK_TAGS = {"p", "li", "blockquote", "dd", "dt", "td", "th"}
    _HEADING_TAGS = {"h1", "h2", "h3", "h4", "h5", "h6"}
    _SKIP_TAGS = {"script", "style", "template", "noscript"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.title: str | None = None
        self.meta_title: str | None = None
        self.published_at_candidates: list[str] = []
        self.chunks: list[IngestedChunk] = []
        self.heading_stack: list[str] = []
        self._heading_only_chunks: list[IngestedChunk] = []
        self._saw_block_content = False
        self._skip_depth = 0
        self._title_capture = False
        self._title_buffer: list[str] = []
        self._heading_capture_tag: str | None = None
        self._heading_buffer: list[str] = []
        self._block_stack: list[str] = []
        self._block_buffer: list[str] = []
        self._heading_paragraph_counts: dict[tuple[str, ...], int] = {}
        self._top_level_paragraph_count = 0

    def _flush_block(self) -> None:
        content = _normalize_whitespace("".join(self._block_buffer))
        self._block_buffer.clear()
        if not content:
            return
        self._saw_block_content = True
        if self.heading_stack:
            key = tuple(self.heading_stack)
            self._heading_paragraph_counts[key] = self._heading_paragraph_counts.get(key, 0) + 1
            location = f"heading:{' > '.join(self.heading_stack)} > paragraph:{self._heading_paragraph_counts[key]}"
        else:
            self._top_level_paragraph_count += 1
            location = f"paragraph:{self._top_level_paragraph_count}"
        self.chunks.append(
            IngestedChunk(
                chunk_id=str(len(self.chunks) + 1),
                location=location,
                content=content,
            )
        )

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        normalized_tag = tag.lower()
        if normalized_tag in self._SKIP_TAGS:
            self._skip_depth += 1
            return
        if self._skip_depth:
            return
        if normalized_tag == "title":
            self._title_capture = True
            self._title_buffer.clear()
            return
        if normalized_tag in self._HEADING_TAGS:
            self._flush_block()
            self._heading_capture_tag = normalized_tag
            self._heading_buffer.clear()
            return
        if normalized_tag == "meta":
            attributes = {key.lower(): value for key, value in attrs if key}
            name = (attributes.get("name") or attributes.get("property") or "").strip().lower()
            content = attributes.get("content") or attributes.get("value")
            if not content:
                return
            if name in _TITLE_META_KEYS and self.meta_title is None:
                self.meta_title = content.strip()
            if name in _PUBLISHED_AT_META_KEYS:
                self.published_at_candidates.append(content)
            return
        if normalized_tag in self._BLOCK_TAGS:
            self._block_stack.append(normalized_tag)

    def handle_endtag(self, tag: str) -> None:
        normalized_tag = tag.lower()
        if normalized_tag in self._SKIP_TAGS:
            self._skip_depth = max(0, self._skip_depth - 1)
            return
        if self._skip_depth:
            return
        if normalized_tag == "title":
            self._title_capture = False
            title = _normalize_whitespace("".join(self._title_buffer))
            if title:
                self.title = title
            self._title_buffer.clear()
            return
        if normalized_tag in self._HEADING_TAGS and self._heading_capture_tag == normalized_tag:
            heading = _normalize_whitespace("".join(self._heading_buffer))
            self._heading_capture_tag = None
            self._heading_buffer.clear()
            if not heading:
                return
            level = int(normalized_tag[1])
            self.heading_stack[:] = self.heading_stack[: level - 1]
            self.heading_stack.append(heading)
            if not self._saw_block_content:
                self._heading_only_chunks.append(
                    IngestedChunk(
                        chunk_id=str(len(self._heading_only_chunks) + 1),
                        location="heading:" + " > ".join(self.heading_stack),
                        content=heading,
                    )
                )
            return
        if normalized_tag in self._BLOCK_TAGS and self._block_stack:
            if self._block_stack[-1] == normalized_tag:
                self._block_stack.pop()
                if not self._block_stack:
                    self._flush_block()
            else:
                while self._block_stack and self._block_stack[-1] != normalized_tag:
                    self._block_stack.pop()
                if self._block_stack and self._block_stack[-1] == normalized_tag:
                    self._block_stack.pop()
                    if not self._block_stack:
                        self._flush_block()

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        if self._title_capture:
            self._title_buffer.append(data)
            return
        if self._heading_capture_tag is not None:
            self._heading_buffer.append(data)
            return
        if self._block_stack:
            self._block_buffer.append(data)

    def close(self) -> None:
        if self._block_stack:
            self._flush_block()
        if not self._saw_block_content and self._heading_only_chunks:
            self.chunks.extend(self._heading_only_chunks)
        super().close()

    def extracted_title(self) -> str | None:
        return self.title or self.meta_title

    def extracted_published_at(self) -> str | None:
        for candidate in self.published_at_candidates:
            normalized = _infer_published_at(candidate)
            if normalized is not None:
                return normalized
        return None


def _parse_saved_page_html(html: str) -> tuple[str | None, str | None, list[IngestedChunk]]:
    parser = _SavedPageParser()
    parser.feed(html)
    parser.close()
    if not parser.chunks:
        if parser.heading_stack:
            parser.chunks.append(
                IngestedChunk(
                    chunk_id="1",
                    location="heading:" + " > ".join(parser.heading_stack),
                    content=parser.heading_stack[-1],
                )
            )
        else:
            raise ValueError("html has no extractable text")
    return parser.extracted_title(), parser.extracted_published_at(), parser.chunks


def _decode_html_bytes(raw: bytes) -> str:
    encodings: list[str] = []
    match = _HTML_CHARSET_RE.search(raw[:4096])
    if match is not None:
        declared = match.group(1).decode("ascii", errors="ignore").strip()
        if declared:
            encodings.append(declared)
    encodings.extend(["utf-8-sig", "utf-8", "cp1252"])

    seen: set[str] = set()
    for encoding in encodings:
        if encoding in seen:
            continue
        seen.add(encoding)
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        except LookupError:
            continue
    return raw.decode("utf-8", errors="replace")


def _read_html_from_path(path: Path, source_type: str) -> str:
    if source_type == "html":
        return _decode_html_bytes(path.read_bytes())
    if source_type != "web-archive":
        raise ValueError(f"unsupported html source type: {source_type}")

    suffix = path.suffix.lower()
    if suffix == ".webarchive":
        with path.open("rb") as fh:
            archive = plistlib.load(fh)
        main_resource = archive.get("WebMainResource") if isinstance(archive, dict) else None
        if not isinstance(main_resource, dict):
            raise ValueError(f"web archive is missing a main HTML resource: {path}")
        resource_data = main_resource.get("WebResourceData")
        if not isinstance(resource_data, (bytes, bytearray)):
            raise ValueError(f"web archive has no HTML resource bytes: {path}")
        encoding = main_resource.get("WebResourceTextEncodingName") or "utf-8"
        try:
            return bytes(resource_data).decode(str(encoding), errors="replace")
        except LookupError:
            return bytes(resource_data).decode("utf-8", errors="replace")

    with path.open("rb") as fh:
        message = message_from_binary_file(fh, policy=policy.default)
    for part in message.walk():
        if part.get_content_type() != "text/html":
            continue
        if part.get_content_disposition() == "attachment":
            continue
        content = part.get_content()
        if isinstance(content, str) and content.strip():
            return content
    raise ValueError(f"web archive has no extractable HTML body: {path}")


def _html_chunks(path: Path, source_type: str) -> tuple[str | None, str | None, list[IngestedChunk]]:
    html = _read_html_from_path(path, source_type)
    return _parse_saved_page_html(html)


def _markdown_chunks(text: str) -> list[IngestedChunk]:
    heading_stack: list[str] = []
    buffer: list[str] = []
    chunks: list[IngestedChunk] = []
    paragraph_index = 0

    def flush() -> None:
        nonlocal paragraph_index
        content = "\n".join(line.rstrip() for line in buffer).strip()
        buffer.clear()
        if not content:
            return
        if heading_stack:
            location = "heading:" + " > ".join(heading_stack)
        else:
            paragraph_index += 1
            location = f"paragraph:{paragraph_index}"
        chunks.append(
            IngestedChunk(
                chunk_id=str(len(chunks) + 1),
                location=location,
                content=content,
            )
        )

    for raw_line in text.splitlines():
        heading_match = re.match(r"^(#{1,6})\s+(.*\S)\s*$", raw_line)
        if heading_match:
            flush()
            level = len(heading_match.group(1))
            heading_text = heading_match.group(2).strip()
            heading_stack[:] = heading_stack[: level - 1]
            heading_stack.append(heading_text)
            continue
        if raw_line.strip() or buffer:
            buffer.append(raw_line)

    flush()
    return chunks


def _csv_chunks(path: Path) -> list[IngestedChunk]:
    chunks: list[IngestedChunk] = []
    with path.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row_index, row in enumerate(reader, start=1):
            parts = [f"{key}={value}" for key, value in row.items() if key is not None]
            chunks.append(
                IngestedChunk(
                    chunk_id=str(row_index),
                    location=f"row:{row_index}",
                    content=", ".join(parts),
                )
            )
    return chunks


def _json_chunks(path: Path) -> list[IngestedChunk]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    chunks: list[IngestedChunk] = []
    if isinstance(payload, list):
        for index, item in enumerate(payload):
            chunks.append(
                IngestedChunk(
                    chunk_id=str(index + 1),
                    location=f"items[{index}]",
                    content=json.dumps(item, sort_keys=True),
                )
            )
        return chunks
    if isinstance(payload, dict):
        for index, key in enumerate(sorted(payload), start=1):
            chunks.append(
                IngestedChunk(
                    chunk_id=str(index),
                    location=f"key:{key}",
                    content=json.dumps(payload[key], sort_keys=True),
                )
            )
        return chunks
    return [
        IngestedChunk(
            chunk_id="1",
            location="root",
            content=json.dumps(payload, sort_keys=True),
        )
    ]


def _pdf_chunks(path: Path) -> list[IngestedChunk]:
    chunks: list[IngestedChunk] = []
    reader = PdfReader(str(path))
    for page_index, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if not text:
            continue
        chunks.append(
            IngestedChunk(
                chunk_id=str(page_index),
                location=f"page:{page_index}",
                content=text,
            )
        )
    if not chunks:
        raise ValueError(f"pdf has no extractable text: {path}")
    return chunks


def _chunks_for_path(path: Path, source_type: str) -> list[IngestedChunk]:
    if source_type == "markdown":
        return _markdown_chunks(path.read_text(encoding="utf-8"))
    if source_type == "csv":
        return _csv_chunks(path)
    if source_type == "json":
        return _json_chunks(path)
    if source_type == "spreadsheet":
        return _parse_spreadsheet_chunks(path)
    if source_type == "pdf":
        return _pdf_chunks(path)
    raise ValueError(f"unsupported source type: {source_type}")


def ingest_file(
    path: Path,
    *,
    source_id: str | None = None,
    title: str | None = None,
    published_at: str | None = None,
    tags: dict[str, str] | None = None,
) -> IngestedDocument:
    source_path = path.resolve()
    source_type = detect_source_type(source_path)
    inferred_title: str | None = None
    inferred_published_at: str | None = None
    chunks: list[IngestedChunk]
    if source_type in {"html", "web-archive"}:
        inferred_title, inferred_published_at, chunks = _html_chunks(source_path, source_type)
    else:
        chunks = _chunks_for_path(source_path, source_type)
    normalized_title = title or inferred_title or source_path.stem
    normalized_published_at = (
        parse_published_at(published_at)
        if published_at is not None
        else inferred_published_at
    )
    normalized_tags = dict(tags or {})
    return IngestedDocument(
        source_id=source_id or _sanitize_source_id(source_path.stem),
        title=normalized_title,
        source_type=source_type,
        path=str(source_path),
        published_at=normalized_published_at,
        tags=normalized_tags,
        chunks=chunks,
    )


def ingest_directory(
    registry: CorpusRegistry,
    path: Path,
    *,
    published_at: str | None = None,
    tags: dict[str, str] | None = None,
) -> dict[str, int]:
    normalized_tags = dict(tags or {})
    summary = {"ingested": 0, "skipped": 0, "failed": 0}
    for candidate in sorted(path.iterdir()):
        if not candidate.is_file():
            continue
        try:
            detect_source_type(candidate)
        except ValueError:
            summary["skipped"] += 1
            continue
        try:
            document = ingest_file(
                candidate,
                published_at=published_at,
                tags=normalized_tags,
            )
            registry.register_document(
                source_id=document.source_id,
                title=document.title,
                source_type=document.source_type,
                path=document.path,
                published_at=document.published_at,
                tags=document.tags,
                chunks=[chunk.__dict__ for chunk in document.chunks],
            )
            summary["ingested"] += 1
        except Exception:
            summary["failed"] += 1
    return summary
